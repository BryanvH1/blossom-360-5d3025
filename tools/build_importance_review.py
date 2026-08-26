# -*- coding: utf-8 -*-
"""Builds one spreadsheet for reviewing the importance weights with John.

    python3 tools/build_importance_review.py

A tab per position: every behaviour, the dimension and lens it belongs to, the weight
we have assigned today, and an empty column to propose a different one. The critical
count is live at the top of each tab, so a swap that breaks the budget of nine is
visible while you are still in the conversation rather than at rebuild time.

Nothing reads this file back — it is a review document. Decisions from it get typed
into the ITEMS tuples in tools/profiles/.
"""
import sys
sys.dont_write_bytecode = True

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from assessment_content import (PROFILES, ROSTER, LENS_DEFS, CRITICAL_BUDGET,
                                OPEN_QS as DEFAULT_OPEN_QS, validate)
from build_assessment import (F_TITLE, F_SUB, F_HDR, F_BODY, F_BOLD, F_SMALL, F_DIM,
                              FILL_HDR, FILL_DIM, FILL_IN, FILL_LT, BOX, WRAP, WRAPT, CTR,
                              NAVY, GREY, BAND_RED, BAND_GRN)

# Bryan keeps the review file in iCloud so it can be opened and shared from anywhere.
# Falls back to the repo's workbooks/ if that folder is not on this machine; override
# either with B360_REVIEW_DIR.
ICLOUD = os.path.expanduser("~/Library/Mobile Documents/com~apple~CloudDocs/MacDocuments/"
                            "blossom/Employees/DeveloperAssessment")
REVIEW_DIR = os.environ.get("B360_REVIEW_DIR") or (
    ICLOUD if os.path.isdir(ICLOUD) else os.path.join(os.path.dirname(HERE), "workbooks"))
OUT = os.path.join(REVIEW_DIR, "ImportanceReview.xlsx")
LENS_NAME = {k: n for k, n, _ in LENS_DEFS}

WEIGHTS = [("3", "Critical", "The role fails without it. Deliberately scarce — "
                             f"{CRITICAL_BUDGET} per position, no more."),
           ("2", "Important", "Expected of the role. Most behaviours sit here."),
           ("1", "Helpful",   "Good to have. Its absence is a nuisance, not a problem.")]


def sheet_name(person):
    """Tab names are capped at 31 characters and cannot contain : \\ / ? * [ ]."""
    n = f"{person['name']} — {person['role']}"
    for ch in ':\\/?*[]':
        n = n.replace(ch, "-")
    return n[:31]


def build():
    errs = validate()
    if errs:
        print("VALIDATION FAILED — not building:")
        for e in errs:
            print("  -", e)
        raise SystemExit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Read me first"
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 22), ("C", 96)]:
        ws.column_dimensions[col].width = w

    ws.cell(row=2, column=2, value="Importance review").font = F_TITLE
    ws.cell(row=3, column=2, value="Blossom 360 · one tab per position · for review with John").font = F_SUB

    intro = [
     ("What this is",
      "Every behaviour we ask about, per position, with the weight we have assigned it today. "
      "Importance says how much THE ROLE needs the behaviour. It is a definition of the job, not "
      "a score of the person, and it is set before any assessment scores arrive."),
     ("Why it matters",
      "Importance is half of the development-plan arithmetic: Gap = Importance × (5 − others' "
      "average). A behaviour weighted 1 has to be scored catastrophically to outrank a critical "
      "scored merely poorly. Change a weight and you change what surfaces as the top five."),
     ("What to do",
      "Read the Behaviour column. If the weight looks wrong, put the weight you would rather see "
      "in the yellow \"Proposed\" column and say why in Notes. Leave it blank where you agree — "
      "blank means no change, so only disagreements need typing."),
     ("The budget",
      f"Exactly {CRITICAL_BUDGET} behaviours per position may be critical. That scarcity is the "
      "point: if everything is critical, nothing is. To promote one to a 3 you have to demote "
      "another. The count at the top of each tab is live and turns red when the proposed set "
      "does not add up."),
     ("Lenses",
      "The Lens column is the second way the same behaviours are grouped — John's four: Skill "
      "Set, Work Ethic, Attitude, and Durability & Improvement. It is there for context while "
      "you weigh things up; nothing needs changing in that column."),
     ("The Open questions tab",
      "The three written questions each rater answers at the end of the form. They carry more "
      "weight in the conversation than any score does — the numbers tell you where to look, these "
      "tell you what is actually going on. Only the Developer form tailors its third question "
      "today; the other four ask a generic version, which is the obvious thing to improve."),
     ("The Metrics tab",
      "One row per behaviour per position: how we would actually measure it, the target, how "
      "often, and where the number comes from. It is there because a weight and a metric are the "
      "same argument — if a behaviour is critical but nobody will ever count it, either the weight "
      "is wrong or the metric is. Worth a skim while you are weighing things up."),
     ("Afterwards",
      "Nothing reads this file back automatically. Send it over and the agreed changes get typed "
      "into the profile definitions, then the forms and workbooks are rebuilt from those."),
    ]
    r = 5
    for label, text in intro:
        c1 = ws.cell(row=r, column=2, value=label); c1.font = F_BOLD; c1.alignment = WRAPT
        c2 = ws.cell(row=r, column=3, value=text);  c2.font = F_BODY; c2.alignment = WRAPT
        ws.row_dimensions[r].height = max(30, 13 * (len(text) // 95 + 1))
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="The three weights").font = F_BOLD
    r += 1
    for val, name, desc in WEIGHTS:
        ws.cell(row=r, column=2, value=f"{val} — {name}").font = F_BOLD
        c = ws.cell(row=r, column=3, value=desc); c.font = F_BODY; c.alignment = WRAPT
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Positions in this file").font = F_BOLD
    r += 1
    for i, h in enumerate(["Tab", "Position", "Behaviours", "Critical", "Important", "Helpful"], start=2):
        c = ws.cell(row=r, column=i, value=h)
        c.font = F_HDR; c.fill = FILL_HDR; c.alignment = CTR; c.border = BOX
    ws.column_dimensions["D"].width = 13
    for col in "EFG":
        ws.column_dimensions[col].width = 11
    summary_hdr = r
    r += 1

    people = [p for p in ROSTER if p["profile"]]
    for person in people:
        prof = PROFILES[person["profile"]]
        items = prof["items"]
        counts = {i: sum(1 for _, _, imp in items if imp == i) for i in (1, 2, 3)}
        ws.cell(row=r, column=2, value=person["name"]).font = F_BODY
        ws.cell(row=r, column=3, value=person["role"]).font = F_BODY
        for col, val in [(4, len(items)), (5, counts[3]), (6, counts[2]), (7, counts[1])]:
            c = ws.cell(row=r, column=col, value=val)
            c.font = F_BODY; c.alignment = CTR; c.border = BOX
        ws.cell(row=r, column=2).border = BOX
        ws.cell(row=r, column=3).border = BOX
        r += 1
    ws.cell(row=r + 1, column=2, value=(
        "The same behaviour can carry a different weight in different positions — that is the "
        "instrument working, not an inconsistency. What a lead developer must do, an associate "
        "may only need to be developing.")).font = F_SMALL
    ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)

    # ------------------------------------------------------------ one tab per role
    for person in people:
        prof = PROFILES[person["profile"]]
        items = prof["items"]
        dim_name = {c: n for c, n, _ in prof["dimensions"]}
        lens_of = {}
        for k, codes in prof["lenses"].items():
            for c in codes:
                lens_of[c] = LENS_NAME[k]

        s = wb.create_sheet(sheet_name(person))
        s.sheet_view.showGridLines = False
        for col, w in [("A", 7), ("B", 26), ("C", 74), ("D", 22), ("E", 11), ("F", 11), ("G", 46)]:
            s.column_dimensions[col].width = w

        s.cell(row=1, column=1, value=f"{person['name']} — {person['role']}").font = F_TITLE
        s.cell(row=2, column=1, value=(
            f"{len(items)} behaviours · 3 = critical, 2 = important, 1 = helpful · "
            f"the budget is {CRITICAL_BUDGET} criticals. Put a different weight in the yellow "
            f"Proposed column only where you disagree.")).font = F_SUB

        FIRST = 6
        LAST = FIRST + len(items) - 1
        s.cell(row=4, column=1, value="Criticals now").font = F_SMALL
        s.cell(row=4, column=2, value=f'=COUNTIF(E{FIRST}:E{LAST},3)').font = F_BOLD
        s.cell(row=4, column=2).alignment = CTR
        s.cell(row=4, column=3, value="Criticals if the proposals are accepted").font = F_SMALL
        s.cell(row=4, column=3).alignment = Alignment(horizontal="right", vertical="center")
        prop = s.cell(row=4, column=4, value=(
            f'=COUNTIF(F{FIRST}:F{LAST},3)+SUMPRODUCT((E{FIRST}:E{LAST}=3)*(F{FIRST}:F{LAST}=""))'))
        prop.font = F_BOLD; prop.alignment = CTR; prop.border = BOX
        s.cell(row=4, column=5, value=f"(budget {CRITICAL_BUDGET})").font = F_SMALL
        s.conditional_formatting.add("D4", CellIsRule(
            operator="notEqual", formula=[str(CRITICAL_BUDGET)],
            fill=PatternFill("solid", bgColor=BAND_RED), font=Font(bold=True, color="8A1C12")))
        s.conditional_formatting.add("D4", CellIsRule(
            operator="equal", formula=[str(CRITICAL_BUDGET)],
            fill=PatternFill("solid", bgColor=BAND_GRN)))

        HR = 5
        for i, h in enumerate(["Code", "Dimension", "Behaviour", "Lens",
                               "Importance", "Proposed", "Notes"], start=1):
            c = s.cell(row=HR, column=i, value=h)
            c.font = F_HDR; c.fill = FILL_HDR; c.border = BOX
            c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        s.row_dimensions[HR].height = 26
        s.freeze_panes = f"A{FIRST}"

        dv = DataValidation(type="whole", operator="between", formula1=1, formula2=3,
                            allow_blank=True, showErrorMessage=True, errorTitle="Weight 1–3",
                            error="Enter 1, 2 or 3 — or leave blank to keep the current weight.")
        s.add_data_validation(dv)

        prev_dim = None
        for i, (code, text, imp) in enumerate(items):
            r = FIRST + i
            s.cell(row=r, column=1, value=code).font = F_SMALL
            s.cell(row=r, column=1).alignment = CTR
            if code[0] != prev_dim:
                prev_dim = code[0]
                s.cell(row=r, column=2, value=dim_name[code[0]]).font = F_BOLD
            s.cell(row=r, column=2).alignment = WRAP
            b = s.cell(row=r, column=3, value=text); b.font = F_BODY; b.alignment = WRAP
            ln = s.cell(row=r, column=4, value=lens_of.get(code, "")); ln.font = F_SMALL; ln.alignment = WRAP
            ci = s.cell(row=r, column=5, value=imp)
            ci.font = F_BOLD if imp == 3 else F_BODY
            ci.alignment = CTR; ci.fill = FILL_LT
            cp = s.cell(row=r, column=6); cp.fill = FILL_IN; cp.alignment = CTR
            dv.add(cp)
            cn = s.cell(row=r, column=7); cn.fill = FILL_IN; cn.alignment = WRAPT
            for col in range(1, 8):
                s.cell(row=r, column=col).border = BOX
            s.row_dimensions[r].height = 30

        s.conditional_formatting.add(f"E{FIRST}:E{LAST}", CellIsRule(
            operator="equal", formula=["3"],
            fill=PatternFill("solid", bgColor="FFD9A0"), font=Font(bold=True)))
        s.conditional_formatting.add(f"F{FIRST}:F{LAST}", CellIsRule(
            operator="equal", formula=["3"],
            fill=PatternFill("solid", bgColor="FFD9A0"), font=Font(bold=True)))

        n = LAST + 2
        s.cell(row=n, column=2, value=(
            "Importance describes what the ROLE needs, not how the person is doing. "
            "Agreeing it before any scores arrive is what keeps it honest — it is much harder to "
            "argue a weight up or down once you can see who it would help.")).font = F_SMALL
        s.merge_cells(start_row=n, start_column=2, end_row=n, end_column=7)
        s.row_dimensions[n].height = 28

    # ------------------------------------------------------- open questions tab
    qs = wb.create_sheet("Open questions")
    qs.sheet_view.showGridLines = False
    for col, w in [("A", 20), ("B", 5), ("C", 78), ("D", 16), ("E", 52), ("F", 40)]:
        qs.column_dimensions[col].width = w

    qs.cell(row=1, column=1, value="The written questions").font = F_TITLE
    qs.cell(row=2, column=1, value=(
        "Three questions at the end of every form, answered in the rater's own words. These carry "
        "more weight in the conversation than any score does — the numbers say where to look, "
        "these say what is actually going on.")).font = F_SUB

    QHR = 4
    for i, h in enumerate(["Position", "#", "Question as asked today", "Written for this role?",
                           "Proposed wording", "Notes"], start=1):
        c = qs.cell(row=QHR, column=i, value=h)
        c.font = F_HDR; c.fill = FILL_HDR; c.border = BOX
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    qs.row_dimensions[QHR].height = 26
    qs.freeze_panes = f"A{QHR + 1}"

    r = QHR + 1
    generic = 0
    for n, person in enumerate(people):
        prof = PROFILES[person["profile"]]
        own = prof["openQuestions"] != DEFAULT_OPEN_QS
        shade = PatternFill("solid", fgColor="F7FAFC") if n % 2 else None
        first_row = r
        for i, q in enumerate(prof["openQuestions"], start=1):
            text = q.split(". ", 1)[1] if q[:2].rstrip(".").isdigit() else q
            tailored = "Yes" if own else "No — generic"
            if not own:
                generic += 1
            for col, v in enumerate([person["name"] if r == first_row else None, i, text,
                                     tailored if r == first_row else None, None, None], start=1):
                c = qs.cell(row=r, column=col, value=v)
                c.font = F_BOLD if col == 1 else F_BODY
                c.alignment = WRAP if col == 3 else CTR
                c.border = BOX
                if shade and col < 5:
                    c.fill = shade
            for col in (5, 6):
                qs.cell(row=r, column=col).fill = FILL_IN
                qs.cell(row=r, column=col).alignment = WRAPT
            qs.cell(row=r, column=1).alignment = WRAPT
            qs.cell(row=r, column=4).alignment = WRAP
            qs.row_dimensions[r].height = 34
            r += 1

    qs.conditional_formatting.add(f"D{QHR + 1}:D{r - 1}", FormulaRule(
        formula=[f'ISNUMBER(SEARCH("generic",D{QHR + 1}))'],
        fill=PatternFill("solid", bgColor="FCE7A2")))

    notes = [
     "Question 3 is the one worth tailoring. Frans's asks what you would need to see before you "
     "would be comfortable with him leading the dev team — a specific decision, so it gets "
     "specific answers. The other four ask the generic version, which gets generic answers.",
     "Keep them to three. Raters answer three questions properly and six of them badly.",
     "They are answered in the rater's own words and attached to their name, like the scores. "
     "The comments get summarised for the person being assessed rather than passed on verbatim.",
    ]
    r += 1
    for note in notes:
        c = qs.cell(row=r, column=1, value=note); c.font = F_SMALL; c.alignment = WRAPT
        qs.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        qs.row_dimensions[r].height = max(26, 13 * (len(note) // 110 + 1))
        r += 1

    # ------------------------------------------------------------- metrics tab
    ms = wb.create_sheet("Metrics")
    ms.sheet_view.showGridLines = False
    for col, w in [("A", 20), ("B", 7), ("C", 60), ("D", 7), ("E", 58),
                   ("F", 22), ("G", 15), ("H", 34), ("I", 32)]:
        ms.column_dimensions[col].width = w

    ms.cell(row=1, column=1, value="Metric library — every behaviour, every position").font = F_TITLE
    ms.cell(row=2, column=1, value=(
        "How each behaviour would actually be counted. A metric nobody will keep is worse than "
        "none — if one here looks unkeepable, say so in Notes and we will swap it for something "
        "you would genuinely tally.")).font = F_SUB

    MHR = 4
    for i, h in enumerate(["Position", "Code", "Behaviour", "Imp.", "Suggested metric",
                           "Target", "Cadence", "How to capture it", "Notes"], start=1):
        c = ms.cell(row=MHR, column=i, value=h)
        c.font = F_HDR; c.fill = FILL_HDR; c.border = BOX
        c.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
    ms.row_dimensions[MHR].height = 26
    ms.freeze_panes = f"C{MHR + 1}"

    r = MHR + 1
    for n, person in enumerate(people):
        prof = PROFILES[person["profile"]]
        shade = PatternFill("solid", fgColor="F7FAFC") if n % 2 else None
        first_row = r
        for code, text, imp in prof["items"]:
            met, tgt, src, cad = prof["metrics"][code]
            vals = [person["name"] if r == first_row else None, code, text, imp,
                    met, tgt, cad, src, None]
            for col, v in enumerate(vals, start=1):
                c = ms.cell(row=r, column=col, value=v)
                c.font = F_BOLD if (col == 1 or (col == 4 and imp == 3)) else F_BODY
                c.alignment = WRAP if col in (3, 5, 8) else CTR
                c.border = BOX
                if shade and col != 9:
                    c.fill = shade
            ms.cell(row=r, column=9).fill = FILL_IN
            ms.cell(row=r, column=1).alignment = WRAPT
            ms.row_dimensions[r].height = 30
            r += 1
        ms.cell(row=first_row, column=1, value=person["name"]).font = F_BOLD

    ms.conditional_formatting.add(f"D{MHR + 1}:D{r - 1}", CellIsRule(
        operator="equal", formula=["3"],
        fill=PatternFill("solid", bgColor="FFD9A0"), font=Font(bold=True)))
    ms.cell(row=r + 1, column=1, value=(
        "Most of these are a count, not a measurement — recurrences, reopens, weeks with a status "
        "post, times feedback had to be repeated. A shared note or one spreadsheet tab, updated "
        "Friday afternoon, is enough. Count events, never adjectives.")).font = F_SMALL
    ms.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    out = build()
    print("saved:", out)
    for p in ROSTER:
        if p["profile"]:
            prof = PROFILES[p["profile"]]
            crit = [c for c, _, i in prof["items"] if i == 3]
            print(f"  {p['name']:8s} {len(prof['items'])} behaviours · criticals {', '.join(crit)}")
