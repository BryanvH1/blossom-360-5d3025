# -*- coding: utf-8 -*-
"""Builds the Blossom 360 workbooks — one per person on the roster.

    python3 tools/build_assessment.py            # rebuild every workbook in workbooks/
    python3 tools/build_assessment.py Frans Mark # rebuild just those

Content comes from assessment_content.py. The rater columns come from
RELATIONSHIP_MAP, so each workbook carries exactly the raters that person
actually has, named rather than numbered, and the direct-report group appears
for the seats that have one. Nothing in here is specific to a single person.
"""
import sys
sys.dont_write_bytecode = True          # no stray __pycache__ in the repo

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.chart import RadarChart, LineChart, Reference

import os

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from assessment_content import (PROFILES, ROSTER, RELATIONSHIPS, RELATIONSHIP_MAP,
                                SCALE, SCALE_NOTES, validate, coverage_warnings)

REPO = os.path.dirname(HERE)
# Workbooks are build artefacts, not source — .gitignore keeps *.xlsx out of the repo.
# Point B360_WORKBOOKS somewhere else (iCloud, a share) to have them land there instead.
OUT_DIR = os.environ.get("B360_WORKBOOKS", os.path.join(REPO, "workbooks"))
REL_LABEL = dict(RELATIONSHIPS)
GROUP_ORDER = [k for k, _ in RELATIONSHIPS]        # self, manager, peer, report

# ---------- palette ----------
INK      = "1F2A37"
NAVY     = "1B3A5C"
ACCENT   = "2E6F8E"
LIGHT    = "EAF1F5"
INPUT    = "FFF7D6"   # yellow input, nod to the original
BAND_RED = "F8C9C4"
BAND_YEL = "FCE7A2"
BAND_GRN = "C9E5C4"
GREY     = "6B7280"

F_TITLE   = Font(name="Calibri", size=16, bold=True, color=NAVY)
F_SUB     = Font(name="Calibri", size=10, italic=True, color=GREY)
F_HDR     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
F_DIM     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_BODY    = Font(name="Calibri", size=10, color=INK)
F_BOLD    = Font(name="Calibri", size=10, bold=True, color=INK)
F_SMALL   = Font(name="Calibri", size=9, color=GREY)

FILL_HDR  = PatternFill("solid", fgColor=ACCENT)
FILL_DIM  = PatternFill("solid", fgColor=NAVY)
FILL_IN   = PatternFill("solid", fgColor=INPUT)
FILL_LT   = PatternFill("solid", fgColor=LIGHT)

thin = Side(style="thin", color="C9D4DC")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

WRAP = Alignment(wrap_text=True, vertical="center")
WRAPT = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------- layout model
def raters_for(name):
    """The rater columns for one person, as (relationship, rater name).

    Self first, then managers, peers and direct reports — exactly the people
    RELATIONSHIP_MAP says rate them. Grouping them in this order keeps each
    group's columns contiguous, which is what lets the group averages stay a
    single AVERAGE() range however many raters a seat has.
    """
    mapped = RELATIONSHIP_MAP.get(name, {})
    roster_order = [r["name"] for r in ROSTER]
    out = [("self", name)]
    for group in GROUP_ORDER[1:]:
        for rater in roster_order:
            if mapped.get(rater) == group:
                out.append((group, rater))
    return out


def dimension_spans(items, first_row):
    """First and last spreadsheet row for each dimension's block of items."""
    spans = {}
    for i, (code, _, _) in enumerate(items):
        d, r = code[0], first_row + i
        if d in spans:
            spans[d][1] = r
        else:
            spans[d] = [r, r]
    for d, (a, b) in spans.items():
        n = sum(1 for c, _, _ in items if c[0] == d)
        if b - a + 1 != n:
            raise SystemExit(f"dimension {d}: items are not contiguous in ITEMS — "
                             f"rows {a}-{b} cover {b - a + 1} of {n}")
    return spans


def filename_for(role):
    """Reproduce the existing workbook names from the role title."""
    return role.replace(" / ", "-").replace(" · ", "-").replace(" ", "_") + "_360.xlsx"


def avg(col_a, col_b, row, nd=2):
    """AVERAGE over a column span on one row, blank when nothing has been entered."""
    a, b = get_column_letter(col_a), get_column_letter(col_b)
    return (f'=IF(COUNT({a}{row}:{b}{row})=0,"",ROUND(AVERAGE({a}{row}:{b}{row}),{nd}))')


def block_avg(sheet, col_a, col_b, row_a, row_b, nd=2):
    """AVERAGE over a rectangle on another sheet, blank when empty."""
    a, b = get_column_letter(col_a), get_column_letter(col_b)
    rng = f"'{sheet}'!{a}{row_a}:{b}{row_b}"
    return f'=IF(COUNT({rng})=0,"",ROUND(AVERAGE({rng}),{nd}))'


# =====================================================================
def build(person):
    name = person["name"]
    role = person["role"]
    profile = PROFILES[person["profile"]]
    DIMENSIONS = profile["dimensions"]
    ITEMS      = profile["items"]
    METRICS    = profile["metrics"]
    OPEN_QS    = profile["openQuestions"]

    raters = raters_for(name)
    n_rat = len(raters)
    if n_rat < 2:
        print(f"  skipped {name}: no raters mapped besides self")
        return None

    # ---- Scores sheet columns, derived from the rater list -------------
    C_CODE, C_BEHAV, C_IMP = 1, 2, 3
    C_FIRST_RATER = 4
    C_SELF = C_FIRST_RATER
    C_LAST_RATER = C_FIRST_RATER + n_rat - 1
    C_FIRST_OTHER = C_SELF + 1                       # others = everyone but self

    groups = []                                      # (key, first_col, last_col)
    col = C_FIRST_RATER
    for key in GROUP_ORDER:
        cols = [c for c, (g, _) in zip(range(C_FIRST_RATER, C_LAST_RATER + 1), raters) if g == key]
        if cols and key != "self":
            groups.append((key, min(cols), max(cols)))
    del col

    # With only one non-self group its average is identical to "Others avg", so the
    # per-group columns only earn their place once there are two or more groups.
    shown = groups if len(groups) > 1 else []

    c = C_LAST_RATER + 1
    C_OTHERS = c; c += 1
    C_GROUP = {}
    for key, _, _ in shown:
        C_GROUP[key] = c; c += 1
    C_BLIND = c; c += 1
    C_GAP   = c; c += 1
    C_RANK  = c; c += 1
    C_SPACER = c; c += 1
    C_KEY   = c; c += 1                              # hidden gap tie-break key
    C_BSKEY = c; c += 1                              # hidden blind-spot key

    FIRST = 4
    LAST = FIRST + len(ITEMS) - 1
    SPANS = dimension_spans(ITEMS, FIRST)

    L = get_column_letter
    wb = openpyxl.Workbook()

    # =====================================================================
    # 1. Instructions
    # =====================================================================
    ws = wb.active
    ws.title = "1 How to Run This"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 96

    def head(ws, row, title, sub=None):
        ws.cell(row=row, column=2, value=title).font = F_TITLE
        if sub:
            ws.cell(row=row + 1, column=2, value=sub).font = F_SUB
        return row + (3 if sub else 2)

    r = head(ws, 2, f"{profile['label']} 360 Assessment",
             f"Blossom · about {name} · {len(ITEMS)} observable behaviours across "
             f"{len(DIMENSIONS)} dimensions")

    rater_line = " · ".join(
        f"{rn} ({REL_LABEL[g].lower()})" for g, rn in raters[1:])

    steps = [
     ("What it does",
      f"{len(ITEMS)} observable behaviours across {len(DIMENSIONS)} dimensions. Everyone scores the same "
      f"{len(ITEMS)}. The workbook then ranks them by IMPORTANCE TO THE ROLE × HOW FAR SHORT THE SCORES "
      "FALL, and hands you the top five development items. You do not pick them — the math does."),
     ("Why it is credible",
      "Each item is a behaviour someone can point at, not a trait. Nobody scores \"attitude\"; they score a "
      "specific thing that either happens or does not. Disagreements become about evidence, which is the "
      "conversation you want."),
     ("The blind-spot column",
      f"Sheet 4 subtracts everyone else's average from {name}'s own self-score. A gap of +1.0 or more means "
      "they think they are fine there and the people around them do not. That column, not the ranking, is "
      "usually what lands in the conversation."),
     ("Step 1 — set importance",
      "On sheet '3 Scores', column C rates each behaviour 1–3 for how much THIS ROLE needs it. Defaults are "
      "filled in. Agree them BEFORE any scores arrive — it is a role definition, not a poll, and setting it "
      "first keeps it honest."),
     ("Step 2 — send the form",
      "Right-click sheet '2 Rater Form' → Move or Copy → New book → save as e.g. "
      "'Assessment - <name>.xlsx' and send it. Ask for it back within a week. "
      f"{name} fills one out on themselves at the same time, before seeing anyone else's."),
     ("Step 3 — enter the scores",
      "Paste each returned column into the matching yellow column on sheet '3 Scores' — the columns are "
      "named, so each rater has one place to go. Blank = not observed; blanks are ignored, not counted as "
      "zero. (The old workbook counted empty rater groups as zero and skewed the averages — this one "
      "does not.)"),
     ("Step 4 — read the results",
      f"Sheet '4 Results' gives the {len(DIMENSIONS)} dimension scores by rater group, the radar chart, and "
      "the blind spots. Sheet '5 Development Plan' auto-fills the top five items with a suggested metric "
      "for each."),
     ("Step 5 — set the metrics",
      "For each of the five items, agree a baseline (what it is today) and a 90-day target. Sheet "
      f"'6 Metric Library' proposes a quantifiable measure for all {len(ITEMS)} behaviours, so whatever "
      "surfaces already has one. Most are a weekly or monthly tally — five minutes a week. Anything "
      "nobody will actually count is not a metric; swap it for one you will."),
     ("Step 6 — track it",
      "Log the readings on sheet '7 Progression' and re-run the full 360 at 90 and 180 days. Two data points "
      "is an anecdote; the trend line is the evidence."),
     ("Who scores",
      f"Self ({name}) · {rater_line}. These come from the roster, not from the raters — nobody picks "
      "their own relationship, because a rater who is a peer one round and a report the next breaks the "
      "comparison. Adding or changing a rater is a one-line edit in RELATIONSHIP_MAP followed by a rebuild."),
     ("One caution",
      f"{n_rat - 1} raters is not anonymous, and everyone will know it. Say so up front rather than "
      f"pretending otherwise — tell them the scores are shared with {name}, the written comments are "
      "summarised, and the point is development, not a file. People score honestly when they know the "
      "rules; they score generously when they are guessing at them."),
    ]
    for label, text in steps:
        c1 = ws.cell(row=r, column=2, value=label); c1.font = F_BOLD; c1.alignment = WRAPT
        c2 = ws.cell(row=r, column=3, value=text);  c2.font = F_BODY; c2.alignment = WRAPT
        ws.row_dimensions[r].height = max(28, 13 * (len(text) // 95 + 1))
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Rating scale").font = F_BOLD
    r += 1
    for val, sname, desc in SCALE:
        ws.cell(row=r, column=2, value=val or "blank").font = F_BOLD
        cc = ws.cell(row=r, column=3, value=sname + " — " + desc)
        cc.font = F_BODY; cc.alignment = WRAPT
        ws.row_dimensions[r].height = max(26, 13 * (len(desc) // 95 + 1))
        r += 1
    r += 1
    for lbl, txt in SCALE_NOTES:
        ws.cell(row=r, column=2, value=lbl).font = F_BOLD
        cc = ws.cell(row=r, column=3, value=txt); cc.font = F_BODY; cc.alignment = WRAPT
        ws.row_dimensions[r].height = max(26, 13 * (len(txt) // 95 + 1))
        r += 1

    # =====================================================================
    # 2. Rater Form
    # =====================================================================
    ws = wb.create_sheet("2 Rater Form")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 8), ("C", 78), ("D", 12), ("E", 4)]:
        ws.column_dimensions[col].width = w

    ws.cell(row=2, column=2, value=f"{profile['label']} 360 — Rater Form").font = F_TITLE
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=4)
    ws.cell(row=3, column=2,
            value=f"About: {name} ({role})   ·   Please return within one week.").font = F_SUB
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=4)

    for i, lbl in enumerate(["Your name:", "Date:"]):
        rr = 5 + i
        ws.cell(row=rr, column=2, value=lbl).font = F_BOLD
        cc = ws.cell(row=rr, column=3); cc.fill = FILL_IN; cc.border = BOX; cc.font = F_BODY

    ws.cell(row=7, column=2, value="Relationship:").font = F_BOLD
    ws.cell(row=7, column=3, value=(
        f"Set by the roster, not chosen here — your column on sheet '3 Scores' already records "
        f"what you are to {name}.")).font = F_SMALL

    r = 9
    ws.cell(row=r, column=2, value=(
        "Score each behaviour 1–5 in the yellow column. Leave it BLANK if you have not seen enough to "
        "judge — a blank is ignored, a guess is not. 5 = exceptional, 4 = strong, 3 = solid/meets the "
        "bar, 2 = developing, 1 = significant development needed.")).font = F_BODY
    ws.cell(row=r, column=2).alignment = WRAPT
    ws.merge_cells(start_row=r, start_column=2, end_row=r + 1, end_column=4)
    ws.row_dimensions[r].height = 15
    ws.row_dimensions[r + 1].height = 15

    dv = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True,
                        showErrorMessage=True, errorTitle="Score 1–5",
                        error="Enter a whole number 1 through 5, or leave blank if you haven't observed it.")
    ws.add_data_validation(dv)

    r = 12
    for dcode, dname, dq in DIMENSIONS:
        cc = ws.cell(row=r, column=2, value=dname)
        cc.font = F_DIM; cc.fill = FILL_DIM
        ws.cell(row=r, column=3).fill = FILL_DIM
        ws.cell(row=r, column=4, value="Score 1–5").font = F_DIM
        ws.cell(row=r, column=4).fill = FILL_DIM
        ws.cell(row=r, column=4).alignment = CTR
        ws.row_dimensions[r].height = 20
        r += 1
        for code, text, imp in ITEMS:
            if not code.startswith(dcode):
                continue
            ws.cell(row=r, column=2, value=code).font = F_SMALL
            cb = ws.cell(row=r, column=3, value=text); cb.font = F_BODY; cb.alignment = WRAP
            cs = ws.cell(row=r, column=4); cs.fill = FILL_IN; cs.border = BOX; cs.alignment = CTR
            dv.add(cs)
            ws.row_dimensions[r].height = 18
            r += 1
        r += 1

    r += 1
    ws.cell(row=r, column=2,
            value="Three questions — please answer in your own words. These matter more than the "
                  "numbers.").font = F_BOLD
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    r += 2
    for q in OPEN_QS:
        ws.cell(row=r, column=2, value=q).font = F_BODY
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
        box = ws.cell(row=r, column=2); box.fill = FILL_IN; box.border = BOX; box.alignment = WRAPT
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 3, end_column=4)
        for k in range(4):
            ws.row_dimensions[r + k].height = 16
        r += 5

    # =====================================================================
    # 3. Scores
    # =====================================================================
    ws = wb.create_sheet("3 Scores")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 68
    ws.column_dimensions["C"].width = 11
    for cc in range(C_FIRST_RATER, C_LAST_RATER + 1):
        ws.column_dimensions[L(cc)].width = 11
    for cc in list(C_GROUP.values()) + [C_OTHERS, C_BLIND, C_GAP]:
        ws.column_dimensions[L(cc)].width = 11
    ws.column_dimensions[L(C_RANK)].width = 7
    ws.column_dimensions[L(C_SPACER)].width = 3
    ws.column_dimensions[L(C_KEY)].hidden = True
    ws.column_dimensions[L(C_BSKEY)].hidden = True

    ws.cell(row=1, column=1, value="Score entry").font = F_TITLE
    ws.cell(row=2, column=1, value=(
        "Yellow = paste rater scores here, one column per named rater. Column C = importance to THIS ROLE "
        "(1 helpful · 2 important · 3 critical), agreed before scores arrive. Blank cells are "
        "ignored, never counted as zero.")).font = F_SUB

    HR = 3
    hdr = {C_CODE: "Code", C_BEHAV: "Behaviour", C_IMP: "Import.",
           C_OTHERS: "Others avg", C_BLIND: "Blind spot", C_GAP: "Gap score", C_RANK: "Rank"}
    for i, (g, rn) in enumerate(raters):
        hdr[C_FIRST_RATER + i] = f"{rn}\n({REL_LABEL[g]})"
    for key, _, _ in shown:
        hdr[C_GROUP[key]] = f"{REL_LABEL[key]} avg"
    for cc, h in hdr.items():
        cell = ws.cell(row=HR, column=cc, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = BOX
    ws.row_dimensions[HR].height = 34
    ws.freeze_panes = f"{L(C_FIRST_RATER)}{FIRST}"

    dv5 = DataValidation(type="whole", operator="between", formula1=1, formula2=5,
                         allow_blank=True, showErrorMessage=True)
    dv3 = DataValidation(type="whole", operator="between", formula1=1, formula2=3,
                         allow_blank=True, showErrorMessage=True)
    ws.add_data_validation(dv5); ws.add_data_validation(dv3)

    SELF_L, OTH_A, OTH_B = L(C_SELF), L(C_FIRST_OTHER), L(C_LAST_RATER)
    OTH_L, BLI_L, GAP_L = L(C_OTHERS), L(C_BLIND), L(C_GAP)
    KEY_L, BSK_L, IMP_L = L(C_KEY), L(C_BSKEY), L(C_IMP)

    for i, (code, text, imp) in enumerate(ITEMS):
        r = FIRST + i
        ws.cell(row=r, column=C_CODE, value=code).font = F_SMALL
        b = ws.cell(row=r, column=C_BEHAV, value=text); b.font = F_BODY; b.alignment = WRAP
        ci = ws.cell(row=r, column=C_IMP, value=imp)
        ci.fill = FILL_IN; ci.border = BOX; ci.alignment = CTR; ci.font = F_BODY
        dv3.add(ci)
        for cc in range(C_FIRST_RATER, C_LAST_RATER + 1):
            cell = ws.cell(row=r, column=cc); cell.fill = FILL_IN; cell.border = BOX; cell.alignment = CTR
            dv5.add(cell)
        ws.cell(row=r, column=C_OTHERS, value=avg(C_FIRST_OTHER, C_LAST_RATER, r)).number_format = "0.00"
        for key, a, bb in shown:
            ws.cell(row=r, column=C_GROUP[key], value=avg(a, bb, r)).number_format = "0.00"
        ws.cell(row=r, column=C_BLIND,
                value=f'=IF(OR({SELF_L}{r}="",{OTH_L}{r}=""),"",ROUND({SELF_L}{r}-{OTH_L}{r},2))'
                ).number_format = "+0.00;-0.00;0.00"
        ws.cell(row=r, column=C_GAP,
                value=f'=IF(OR({OTH_L}{r}="",{IMP_L}{r}=""),"",ROUND({IMP_L}{r}*(5-{OTH_L}{r}),2))'
                ).number_format = "0.00"
        ws.cell(row=r, column=C_RANK,
                value=f'=IF({KEY_L}{r}="","",RANK({KEY_L}{r},${KEY_L}${FIRST}:${KEY_L}${LAST},0))'
                ).number_format = "0"
        # Hidden rank key. Ties on gap are common with few raters, so the tie-break is explicit:
        # gap -> importance -> blind spot -> item order. Must match assets/core.js on the web version.
        ws.cell(row=r, column=C_KEY, value=(
            f'=IF({GAP_L}{r}="","",ROUND({GAP_L}{r}*100,0)*1000000+{IMP_L}{r}*100000'
            f'+ROUND(MAX(0,IF({BLI_L}{r}="",0,{BLI_L}{r}))*100,0)*100+({LAST + 1}-ROW()))'))
        ws.cell(row=r, column=C_BSKEY,
                value=f'=IF({BLI_L}{r}="","",{BLI_L}{r}-ROW()*0.00001)')
        for cc in list(C_GROUP.values()) + [C_OTHERS, C_BLIND, C_GAP, C_RANK]:
            cell = ws.cell(row=r, column=cc); cell.font = F_BODY; cell.alignment = CTR; cell.border = BOX
            if cc == C_OTHERS or cc in C_GROUP.values():
                cell.fill = FILL_LT

    rng = f"{OTH_L}{FIRST}:{OTH_L}{LAST}"
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["3"],
                                                  fill=PatternFill("solid", bgColor=BAND_RED)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["3", "3.99"],
                                                  fill=PatternFill("solid", bgColor=BAND_YEL)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["4"],
                                                  fill=PatternFill("solid", bgColor=BAND_GRN)))
    ws.conditional_formatting.add(f"{BLI_L}{FIRST}:{BLI_L}{LAST}",
        CellIsRule(operator="greaterThanOrEqual", formula=["1"],
                   fill=PatternFill("solid", bgColor=BAND_RED), font=Font(bold=True, color="8A1C12")))
    ws.conditional_formatting.add(f"{GAP_L}{FIRST}:{GAP_L}{LAST}",
        ColorScaleRule(start_type="min", start_color="FFFFFF", end_type="max", end_color="E8927C"))
    ws.conditional_formatting.add(f"{L(C_RANK)}{FIRST}:{L(C_RANK)}{LAST}",
        CellIsRule(operator="lessThanOrEqual", formula=["5"],
                   fill=PatternFill("solid", bgColor="FFD9A0"), font=Font(bold=True)))

    r = LAST + 2
    ws.cell(row=r, column=2, value=(
        "Gap score = Importance × (5 − Others' average). Highest gaps become the development plan. "
        "Ties break by importance, then blind spot, then item order — if several items tie at the fifth "
        "slot, treat which one makes the cut as a judgement call, not a result.")).font = F_SMALL
    ws.cell(row=r + 1, column=2, value=(
        f"Blind spot = Self − Others. +1.00 or more (shaded red) = {name} rates themselves materially "
        "higher than the people around them do.")).font = F_SMALL
    ws.cell(row=r + 2, column=2, value=(
        f"'Others' deliberately excludes {name}'s own self-score, so they cannot lift their own result."
        )).font = F_SMALL

    r += 4
    ws.cell(row=r, column=2, value="Written comments — paste each rater's answers below").font = F_BOLD
    r += 1
    for q in OPEN_QS:
        ws.cell(row=r, column=2, value=q).font = F_BOLD
        r += 1
        for g, rn in raters:
            ws.cell(row=r, column=1, value=f"{rn} ({REL_LABEL[g]})").font = F_SMALL
            cc = ws.cell(row=r, column=2); cc.fill = FILL_IN; cc.border = BOX; cc.alignment = WRAPT
            r += 1
        r += 1

    # =====================================================================
    # 4. Results
    # =====================================================================
    ws = wb.create_sheet("4 Results")
    ws.sheet_view.showGridLines = False

    R_DIM = 2
    R_SELF = 3
    R_GRP = {}
    c = 4
    for key, _, _ in shown:
        R_GRP[key] = c; c += 1
    R_OTHERS = c; c += 1
    R_PCT = c; c += 1
    R_BLIND = c; c += 1
    R_WHAT = c

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions[L(R_DIM)].width = 34
    for cc in [R_SELF] + list(R_GRP.values()):
        ws.column_dimensions[L(cc)].width = 12
    ws.column_dimensions[L(R_OTHERS)].width = 13
    ws.column_dimensions[L(R_PCT)].width = 11
    ws.column_dimensions[L(R_BLIND)].width = 13
    ws.column_dimensions[L(R_WHAT)].width = 44

    ws.cell(row=2, column=2, value="Results").font = F_TITLE
    ws.cell(row=3, column=2, value=("Scores are averages on the 1–5 scale. Red under 3.0 · amber "
                                    "3.0–3.9 · green 4.0 and up.")).font = F_SUB

    HR = 5
    heads = {R_DIM: "Dimension", R_SELF: "Self", R_OTHERS: "Others avg",
             R_PCT: "% of 5", R_BLIND: "Blind spot", R_WHAT: "What it answers"}
    for key, _, _ in shown:
        heads[R_GRP[key]] = REL_LABEL[key]
    for cc, h in heads.items():
        cell = ws.cell(row=HR, column=cc, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CTR; cell.border = BOX
    ws.row_dimensions[HR].height = 22

    r = HR + 1
    for dcode, dname, dq in DIMENSIONS:
        s, e = SPANS[dcode]
        ws.cell(row=r, column=R_DIM, value=dname).font = F_BOLD
        ws.cell(row=r, column=R_SELF, value=block_avg("3 Scores", C_SELF, C_SELF, s, e))
        for key, a, bb in shown:
            ws.cell(row=r, column=R_GRP[key], value=block_avg("3 Scores", a, bb, s, e))
        ws.cell(row=r, column=R_OTHERS, value=block_avg("3 Scores", C_FIRST_OTHER, C_LAST_RATER, s, e))
        ws.cell(row=r, column=R_PCT,
                value=f'=IF({L(R_OTHERS)}{r}="","",{L(R_OTHERS)}{r}/5)')
        ws.cell(row=r, column=R_BLIND, value=(
            f'=IF(OR({L(R_SELF)}{r}="",{L(R_OTHERS)}{r}=""),"",'
            f'ROUND({L(R_SELF)}{r}-{L(R_OTHERS)}{r},2))'))
        ws.cell(row=r, column=R_WHAT, value=dq).font = F_SMALL
        ws.cell(row=r, column=R_WHAT).alignment = WRAP
        for cc in range(R_SELF, R_WHAT):
            cell = ws.cell(row=r, column=cc); cell.alignment = CTR; cell.border = BOX; cell.font = F_BODY
            cell.number_format = ("0%" if cc == R_PCT else
                                  "+0.00;-0.00;0.00" if cc == R_BLIND else "0.00")
        ws.row_dimensions[r].height = 22
        r += 1

    TOT = r
    ws.cell(row=TOT, column=R_DIM, value="OVERALL").font = F_BOLD
    for cc in [R_SELF] + list(R_GRP.values()) + [R_OTHERS]:
        letter = L(cc)
        ws.cell(row=TOT, column=cc, value=(
            f'=IF(COUNT({letter}{HR + 1}:{letter}{TOT - 1})=0,"",'
            f'ROUND(AVERAGE({letter}{HR + 1}:{letter}{TOT - 1}),2))'))
    ws.cell(row=TOT, column=R_PCT, value=f'=IF({L(R_OTHERS)}{TOT}="","",{L(R_OTHERS)}{TOT}/5)')
    ws.cell(row=TOT, column=R_BLIND, value=(
        f'=IF(OR({L(R_SELF)}{TOT}="",{L(R_OTHERS)}{TOT}=""),"",'
        f'ROUND({L(R_SELF)}{TOT}-{L(R_OTHERS)}{TOT},2))'))
    for cc in range(R_DIM, R_WHAT):
        cell = ws.cell(row=TOT, column=cc); cell.fill = FILL_LT; cell.border = BOX; cell.font = F_BOLD
        if cc > R_DIM:
            cell.alignment = CTR
            cell.number_format = ("0%" if cc == R_PCT else
                                  "+0.00;-0.00;0.00" if cc == R_BLIND else "0.00")

    band = f"{L(R_SELF)}{HR + 1}:{L(R_OTHERS)}{TOT}"
    ws.conditional_formatting.add(band, CellIsRule(operator="lessThan", formula=["3"],
                                                   fill=PatternFill("solid", bgColor=BAND_RED)))
    ws.conditional_formatting.add(band, CellIsRule(operator="between", formula=["3", "3.99"],
                                                   fill=PatternFill("solid", bgColor=BAND_YEL)))
    ws.conditional_formatting.add(band, CellIsRule(operator="greaterThanOrEqual", formula=["4"],
                                                   fill=PatternFill("solid", bgColor=BAND_GRN)))
    ws.conditional_formatting.add(f"{L(R_BLIND)}{HR + 1}:{L(R_BLIND)}{TOT}",
        CellIsRule(operator="greaterThanOrEqual", formula=["0.75"],
                   fill=PatternFill("solid", bgColor=BAND_RED), font=Font(bold=True, color="8A1C12")))

    chart = RadarChart()
    chart.type = "marker"
    chart.style = 26
    chart.title = ("Self vs. " + " vs. ".join(REL_LABEL[k] for k, _, _ in shown)
                   if shown else "Self vs. Others")
    data = Reference(ws, min_col=R_SELF, max_col=R_SELF + len(shown), min_row=HR, max_row=TOT - 1)
    cats = Reference(ws, min_col=R_DIM, min_row=HR + 1, max_row=TOT - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 11; chart.width = 15
    ws.add_chart(chart, f"{L(R_DIM)}{TOT + 3}")

    r = TOT + 26
    ws.cell(row=r, column=2, value=(
        f"Biggest blind spots — where {name} rates themselves highest above everyone else")).font = F_BOLD
    r += 1
    for i, h in enumerate(["Rank", "Code", "Behaviour", "Self", "Others", "Blind spot"], start=2):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CTR; cell.border = BOX
    BS_HDR = r
    for k in range(1, 6):
        rr = r + k
        ws.cell(row=rr, column=2, value=k).alignment = CTR
        mk = (f"MATCH(LARGE('3 Scores'!${BSK_L}${FIRST}:${BSK_L}${LAST},{k}),"
              f"'3 Scores'!${BSK_L}${FIRST}:${BSK_L}${LAST},0)")
        srcs = [L(C_CODE), L(C_BEHAV), SELF_L, OTH_L, BLI_L]
        for col, src in zip(range(3, 8), srcs):
            ws.cell(row=rr, column=col,
                    value=f"=IFERROR(INDEX('3 Scores'!${src}${FIRST}:${src}${LAST},{mk}),\"\")")
        ws.cell(row=rr, column=4).alignment = WRAP
        ws.cell(row=rr, column=4).font = F_BODY
        for col in [2, 3, 5, 6, 7]:
            ws.cell(row=rr, column=col).alignment = CTR
            ws.cell(row=rr, column=col).font = F_BODY
        for col in range(2, 8):
            ws.cell(row=rr, column=col).border = BOX
        for col in (5, 6, 7):
            ws.cell(row=rr, column=col).number_format = "0.00" if col < 7 else "+0.00;-0.00;0.00"
        ws.row_dimensions[rr].height = 22
    ws.cell(row=BS_HDR + 7, column=2, value=(
        "A blind spot of +1.00 or more is worth raising even if the absolute score is fine — the "
        "disagreement about reality is the problem.")).font = F_SMALL

    # =====================================================================
    # 5. Development Plan
    # =====================================================================
    ws = wb.create_sheet("5 Development Plan")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 5), ("C", 7), ("D", 50), ("E", 9), ("F", 9), ("G", 9),
                   ("H", 46), ("I", 20), ("J", 14), ("K", 16), ("L", 13), ("M", 16)]:
        ws.column_dimensions[col].width = w

    ws.cell(row=2, column=2, value="Development Plan — the five items").font = F_TITLE
    ws.cell(row=3, column=2, value=(
        f"Auto-selected by gap score from sheet 3. Nobody chose these; they fell out of the data. "
        f"Fill the yellow cells with {name} in the room.")).font = F_SUB

    HR = 5
    heads = ["#", "Code", "Behaviour", "Self", "Others", "Gap", "Suggested metric", "Target",
             "Baseline (today)", "90-day target", "Cadence", "Who counts it"]
    for i, h in enumerate(heads, start=2):
        cell = ws.cell(row=HR, column=i, value=h); cell.font = F_HDR; cell.fill = FILL_HDR
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = BOX
    ws.row_dimensions[HR].height = 30

    MET_LAST = 3 + len(ITEMS)
    RANK_L = L(C_RANK)
    for k in range(1, 6):
        r = HR + k
        m = f"MATCH({k},'3 Scores'!${RANK_L}${FIRST}:${RANK_L}${LAST},0)"
        ws.cell(row=r, column=2, value=k).alignment = CTR
        for col, src in zip(range(3, 8), [L(C_CODE), L(C_BEHAV), SELF_L, OTH_L, GAP_L]):
            ws.cell(row=r, column=col,
                    value=f"=IFERROR(INDEX('3 Scores'!${src}${FIRST}:${src}${LAST},{m}),\"\")")
        ws.cell(row=r, column=8,
                value=f'=IFERROR(VLOOKUP($C{r},\'6 Metric Library\'!$A$4:$E${MET_LAST},3,FALSE),"")')
        ws.cell(row=r, column=9,
                value=f'=IFERROR(VLOOKUP($C{r},\'6 Metric Library\'!$A$4:$E${MET_LAST},4,FALSE),"")')
        for col in (10, 11, 12, 13):
            cell = ws.cell(row=r, column=col)
            cell.fill = FILL_IN; cell.border = BOX; cell.alignment = WRAP; cell.font = F_BODY
        ws.cell(row=r, column=12,
                value=f'=IFERROR(VLOOKUP($C{r},\'6 Metric Library\'!$A$4:$E${MET_LAST},5,FALSE),"")')
        ws.cell(row=r, column=12).fill = FILL_IN
        for col in range(2, 14):
            cell = ws.cell(row=r, column=col); cell.border = BOX
            if cell.font is None or cell.font.name is None:
                cell.font = F_BODY
        for col in (4, 9, 10):
            ws.cell(row=r, column=col).alignment = WRAP
            ws.cell(row=r, column=col).font = F_BODY
        for col in (2, 3, 5, 6, 7):
            ws.cell(row=r, column=col).alignment = CTR
            ws.cell(row=r, column=col).font = F_BODY
        for col in (5, 6, 7):
            ws.cell(row=r, column=col).number_format = "0.00"
        ws.row_dimensions[r].height = 58

    r = HR + 8
    notes = [
     ("Rules that make this work", ""),
     ("Five, not fifteen",
      "Anything past five stops being a plan and becomes a performance file. If a sixth matters more "
      "later, retire one."),
     ("Baseline before target",
      "Measure two to four weeks of the current reality before setting the target. A target set from a "
      "guess gets argued about instead of hit."),
     (f"{name} keeps the log",
      f"{name} records the readings; whoever runs the check-in verifies them monthly. Ownership of the "
      "measurement is itself part of the assessment — someone who won't track it is telling you "
      "something."),
     ("Monthly, 30 minutes",
      "Read the numbers, not the impressions. \"Three surprises last month, zero this month\" ends an "
      "argument that \"communication is better\" cannot."),
     ("Re-run the 360 at 90 and 180 days",
      f"Same {len(ITEMS)} items, same raters. Movement in the scores is the outcome; movement in the "
      "metrics is the leading indicator that gets you there."),
     ("Name the bar before they start",
      "Agree up front what result counts as done — e.g. all five items at target and no dimension "
      "under 3.5 at the 180-day re-run. Naming it first is the difference between a development plan and "
      "a moving goalpost."),
    ]
    for lbl, txt in notes:
        ws.cell(row=r, column=2, value=lbl).font = F_BOLD if txt else F_TITLE
        if txt:
            cell = ws.cell(row=r, column=4, value=txt); cell.font = F_BODY; cell.alignment = WRAPT
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)
            ws.row_dimensions[r].height = max(26, 13 * (len(txt) // 110 + 1))
        r += 1

    # =====================================================================
    # 6. Metric Library
    # =====================================================================
    ws = wb.create_sheet("6 Metric Library")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 7), ("B", 56), ("C", 60), ("D", 28), ("E", 18), ("F", 40)]:
        ws.column_dimensions[col].width = w

    ws.cell(row=1, column=1, value="Metric Library").font = F_TITLE
    ws.cell(row=2, column=1, value=(
        f"A quantifiable measure for all {len(ITEMS)} behaviours, so whatever the assessment surfaces "
        "already has one. Sheet 5 pulls from here automatically. Edit freely — a metric nobody will "
        "actually count is worse than none.")).font = F_SUB

    for i, h in enumerate(["Code", "Behaviour", "Suggested quantifiable metric", "Suggested target",
                           "Cadence", "How to capture it"], start=1):
        cell = ws.cell(row=3, column=i, value=h); cell.font = F_HDR; cell.fill = FILL_HDR
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = BOX
    ws.row_dimensions[3].height = 24

    for i, (code, text, imp) in enumerate(ITEMS):
        r = 4 + i
        met, tgt, src, cad = METRICS[code]
        for col, val in [(1, code), (2, text), (3, met), (4, tgt), (5, cad), (6, src)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = F_BODY; cell.alignment = WRAP if col in (2, 3, 6) else CTR; cell.border = BOX
        if i % 6 < 3:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F7FAFC")
        ws.row_dimensions[r].height = 32

    r = MET_LAST + 3
    ws.cell(row=r, column=1, value="Keeping the tally").font = F_BOLD
    notes2 = [
     "Most of these are a count, not a measurement — surprises, reopens, weeks with a status post, "
     "times feedback had to be repeated. A shared note or a single spreadsheet tab, updated Friday "
     "afternoon, is enough.",
     "Anything marked Linear, ClickUp or GitHub can be automated later if you want it; none of it needs "
     "to be to start. Start with the tally, automate what survives.",
     "Count events, never adjectives. \"Two items reopened\" is checkable. \"Quality improved\" is a debate.",
     "If a metric goes three weeks without being filled in, that is the answer about whether it was the "
     "right metric. Replace it.",
    ]
    for note in notes2:
        r += 1
        cell = ws.cell(row=r, column=1, value=note); cell.font = F_BODY; cell.alignment = WRAPT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        ws.row_dimensions[r].height = 28

    # =====================================================================
    # 7. Progression
    # =====================================================================
    ws = wb.create_sheet("7 Progression")
    ws.sheet_view.showGridLines = False
    for col, w in [("A", 3), ("B", 34), ("C", 14), ("D", 14), ("E", 14), ("F", 14), ("G", 14)]:
        ws.column_dimensions[col].width = w

    ws.cell(row=2, column=2, value="Progression").font = F_TITLE
    ws.cell(row=3, column=2, value=(
        "Re-run the 360 at 90 and 180 days and record the dimension scores here. Log the five metric "
        "readings below that. Two points is an anecdote — the line is the evidence.")).font = F_SUB

    HR = 5
    cols = ["Baseline", "+90 days", "+180 days", "+270 days", "+1 year"]
    ws.cell(row=HR, column=2, value="Others' average by dimension").font = F_HDR
    ws.cell(row=HR, column=2).fill = FILL_HDR
    for i, cn in enumerate(cols, start=3):
        cell = ws.cell(row=HR, column=i, value=cn)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CTR; cell.border = BOX
    ws.row_dimensions[HR].height = 22

    RES_OTH = L(R_OTHERS)
    RES_FIRST = 6                                    # '4 Results' HR + 1
    r = HR + 1
    for idx, (dcode, dname, dq) in enumerate(DIMENSIONS):
        ws.cell(row=r, column=2, value=dname).font = F_BODY
        ws.cell(row=r, column=3, value=f"='4 Results'!{RES_OTH}{RES_FIRST + idx}").number_format = "0.00"
        ws.cell(row=r, column=3).fill = FILL_LT
        for col in range(4, 8):
            cell = ws.cell(row=r, column=col)
            cell.fill = FILL_IN; cell.number_format = "0.00"; cell.alignment = CTR
        for col in range(2, 8):
            ws.cell(row=r, column=col).border = BOX
        r += 1
    ws.cell(row=r, column=2, value="Overall").font = F_BOLD
    ws.cell(row=r, column=3,
            value=f"='4 Results'!{RES_OTH}{RES_FIRST + len(DIMENSIONS)}").number_format = "0.00"
    for col in range(3, 8):
        cell = ws.cell(row=r, column=col); cell.border = BOX; cell.alignment = CTR
        cell.number_format = "0.00"
        cell.fill = FILL_LT if col == 3 else FILL_IN
        cell.font = F_BOLD
    LAST_DIM = r

    lc = LineChart()
    lc.title = "Dimension trend — others' average"
    lc.y_axis.title = "Score (1–5)"
    lc.style = 12
    d = Reference(ws, min_col=2, max_col=7, min_row=HR + 1, max_row=LAST_DIM - 1)
    lc.add_data(d, titles_from_data=True, from_rows=True)
    lc.set_categories(Reference(ws, min_col=3, max_col=7, min_row=HR))
    lc.height = 9; lc.width = 18
    ws.add_chart(lc, f"B{LAST_DIM + 2}")

    r = LAST_DIM + 22
    ws.cell(row=r, column=2, value="Metric readings — the five development items").font = F_BOLD
    r += 1
    for i, h in enumerate(["Item / metric", "Baseline", "Month 1", "Month 2", "Month 3", "Month 4",
                           "Month 5", "Month 6", "Target", "On track?"], start=2):
        cell = ws.cell(row=r, column=i, value=h)
        cell.font = F_HDR; cell.fill = FILL_HDR; cell.alignment = CTR; cell.border = BOX
    for k in range(1, 6):
        rr = r + k
        ws.cell(row=rr, column=2, value=(
            f'=IFERROR(\'5 Development Plan\'!C{5 + k}&" — "&\'5 Development Plan\'!H{5 + k},"")'
            )).font = F_BODY
        ws.cell(row=rr, column=2).alignment = WRAP
        for col in range(3, 12):
            cell = ws.cell(row=rr, column=col)
            cell.fill = FILL_IN; cell.border = BOX; cell.alignment = CTR; cell.font = F_BODY
        ws.row_dimensions[rr].height = 34
    for w in ["H", "I", "J", "K", "L"]:
        ws.column_dimensions[w].width = 12
    ws.column_dimensions["B"].width = 52

    out = os.path.join(OUT_DIR, filename_for(role))
    wb.save(out)
    return out, raters


# =====================================================================
if __name__ == "__main__":
    errs = validate()
    if errs:
        print("VALIDATION FAILED — not building:")
        for e in errs:
            print("  -", e)
        raise SystemExit(1)
    for w in coverage_warnings():
        print("note:", w)

    wanted = sys.argv[1:]
    known = {r["name"] for r in ROSTER}
    for w in wanted:
        if w not in known:
            raise SystemExit(f"unknown person '{w}' — roster is: {', '.join(sorted(known))}")

    os.makedirs(OUT_DIR, exist_ok=True)
    built = 0
    for person in ROSTER:
        if wanted and person["name"] not in wanted:
            continue
        if not person["profile"]:
            print(f"  skipped {person['name']}: no profile defined")
            continue
        result = build(person)
        if not result:
            continue
        out, raters = result
        cols = ", ".join(f"{rn}/{g}" for g, rn in raters)
        print(f"  {person['name']:7s} -> {os.path.basename(out)}  [{cols}]")
        built += 1
    print(f"{built} workbook(s) written to {OUT_DIR}")
